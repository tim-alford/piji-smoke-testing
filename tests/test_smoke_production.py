# vim: ts=2
import os
import json
from time import sleep
import unittest
from selenium import webdriver
from selenium.webdriver import ChromeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from pandas import read_excel
import openpyxl

class SmokeTestProduction(unittest.TestCase):
		
	@classmethod
	def loadTerms(cls):
		with open("./data/terms.json", "r") as f:
			return json.load(f)
			
	@classmethod
	def setUpClass(cls):
		cls.options = ChromeOptions()
		cls.options.add_experimental_option("prefs", {
			"download.default_directory": ".",
			"download.prompt_for_download": False,
			"download.directory_upgrade": True,
			"safebrowsing.enabled": True
		})
		cls.driver = webdriver.Chrome(options=cls.options)
		if "WEBSITE_FOR_TESTING" not in os.environ.keys():
			raise Exception("Please set the WEBSITE_FOR_TESTING variable before running the suite")
		cls.website = os.environ["WEBSITE_FOR_TESTING"]
		cls.businessWebsite = f"{os.environ['WEBSITE_FOR_TESTING']}/businesses"
		if "TEST_ENV" not in os.environ.keys():
			raise Exception("Please set the TEST_ENV variable before running the suite")
		cls.env = os.environ["TEST_ENV"]
		cls.terms = cls.loadTerms()[cls.env]
		cls.ORGANISATION_TYPES = [x.strip().lower() for x in ["Government", "Government agency", "Industry group", "Informal group", "Lobby group", "Non-government organisation", "Sector peak body", "Other"]]
		
	@classmethod
	def tearDownClass(cls):
		cls.driver.quit()
	
	def does_download_exist(self, fileName):
		user = os.environ["USER"]
		downloadPath = f"/home/{user}/Downloads/{fileName}"
		print(f"Checking path {downloadPath}")
		return (os.path.exists(downloadPath), downloadPath)
		
	def test_business_export_is_working(self):
		cls = self.__class__
		cls.driver.get(cls.website)
		outlets = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "OutletTable"))
		businesses = cls.driver.find_element(By.ID, "businessPage")
		self.assertTrue(businesses is not None)
		businesses.click()
		entities = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "BusinessTable"))
		cls.driver.find_element(By.ID, "DataMenu").click()
		cls.driver.find_element(By.ID, "downloadExport").click()
		download = "Australian News Index - News Producers - PIJI.csv"
		destination = None
		timeout = 60
		while True:
			(result, path) = self.does_download_exist(download)
			if result:
					destination = path
					break
			sleep(1)
			timeout -= 1
			if timeout <= 0:
				self.fail("Failed to find export within 60 seconds")
				break
		self.assertTrue(destination is not None, "Failed to find CSV export.")
		with open(destination ,"r") as f:
			records = [x.strip().split("|") for x in f.readlines()]
		os.remove(destination)
		headers = records[0]
		self.assertTrue("entity_name" in headers)
		self.assertTrue("entity_type" in headers)
		self.assertTrue("entity_abn" in headers)
		self.assertTrue("parent_entities" in headers)
		self.assertTrue("child_entities" in headers)
		self.assertTrue("associated_outlets" in headers)
		records = [records[x] for x in range(0, len(records)) if x > 0]
		entityNameIndex = headers.index("entity_name")
		entityTypeIndex = headers.index("entity_type")
		entityABNIndex = headers.index("entity_abn")
		checked = 0
		for r in records:
			name = r[entityNameIndex]
			abn = r[entityABNIndex]
			entityType = r[entityTypeIndex]
			name = name.strip().strip("\"")
			abn = abn.strip().strip("\"").strip().upper()
			entityType = entityType.strip().strip("\"").strip().upper()
			self.search_for(name)
			searched = name.strip().upper()
			businesses = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "BusinessTable"))
			ids = self.get_business_ids(businesses)
			self.assertTrue(len(ids) > 0, "There should be at least one matching row")
			first = ids[0]
			nameSelector = f"GenericTableCell_Name_{first}"
			abnSelector = f"GenericTableCell_ABN_{first}"
			entityTypeSelector = f"GenericTableCell_Entity type_{first}"
			nameCell = cls.driver.find_element(By.ID, nameSelector)
			abnCell = cls.driver.find_element(By.ID, abnSelector)
			entityTypeCell = cls.driver.find_element(By.ID, entityTypeSelector)
			self.assertTrue(nameCell is not None)
			self.assertTrue(abnCell is not None)
			self.assertTrue(entityTypeCell is not None)
			nameValue = nameCell.text.strip().upper()
			abnValue = abnCell.text.strip().upper()
			entityTypeValue = entityTypeCell.text.strip().upper()
			self.assertTrue(nameValue.find(searched) != -1, f"The search value {searched} was not equal to {nameValue}")
			self.assertTrue(abnValue.find(abn) != -1, f"The search value {abn} was not equal to {abnValue}")
			self.assertTrue(entityTypeValue.find(entityType) != -1, f"The search value {entityType} was not equal to {entityTypeValue}")
			self.clear_search()
			checked += 1
			if checked == 60:
				break

	def test_outlet_export_is_working(self):
		cls = self.__class__
		cls.driver.get(cls.website)
		outlets = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "OutletTable"))
		cls.driver.find_element(By.ID, "DataMenu").click()
		cls.driver.find_element(By.ID, "downloadExport").click()
		download = "Australian News Index (PIJI).xlsx"
		destination = None
		timeout = 60
		while True:
			(result, path) = self.does_download_exist(download)
			if result:
					destination = path
					break
			sleep(1)
			timeout -= 1
			if timeout <= 0:
				self.fail("Failed to find export within 60 seconds")
				break
		self.assertTrue(destination is not None, "Failed to find XLSX export.")
		cls.driver.save_screenshot("./screenshots/test_outlet_export_is_working.png")
		try:
			book = openpyxl.load_workbook(destination)
			self.assertEqual(len(book.sheetnames), 2)
			self.assertTrue("News producers" in book.sheetnames)
			self.assertTrue("News entities" in book.sheetnames)
			outlets = book["News producers"]
			entities = book["News entities"]
			self.assertTrue(outlets.max_row > 1, "Number of outlets in export should be non zero")
			self.assertTrue(entities.max_row > 1, "Number of entities in export should be non zero")
		finally:
			os.remove(destination)
	
	def get_business_ids(self, table):
		rows = table.find_elements(By.TAG_NAME, "tr")
		self.assertTrue(len(rows) > 0, "Business count should be non zero")
		ids = []
		# find all outlet ids on this page
		for row in rows:
			id = row.get_attribute("id")
			if id is not None and id != "":
				tokens = id.split("_")
				entityId = tokens[-1]
				firstPart = tokens[0]
				if firstPart != "GenericTableRow":
					continue
				ids.append(entityId)
		return ids

	def get_outlet_ids(self, table):
		cls = self.__class__
		rows = cls.driver.find_elements(By.TAG_NAME, "tr")
		self.assertTrue(len(rows) > 0, "Outlet count should be non zero")
		ids = []
		# find all outlet ids on this page
		for row in rows:
			try:
				id = row.get_attribute("id")
			except Exception:
				continue
			if id is not None and id != "":
				tokens = id.split("_")
				if tokens[0] != "Outlet":
					continue
				outletId = tokens[-1]
				ids.append(outletId)
		return ids

	def test_search_outlets_by_name(self):
		cls = self.__class__
		cls.driver.get(cls.website)
		outlets = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "OutletTable"))
		ids = self.get_outlet_ids(outlets)
		term = cls.terms["search_outlets"].lower()
		self.search_for(term)
		outlets = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "OutletTable"))
		ids = self.get_outlet_ids(outlets)
		for i in ids:
			name = f"Outlet_{i}_name"
			name = cls.driver.find_element(By.ID, name)
			name = name.text.strip().lower()
			self.assertTrue(term in name, f"Failed to find term {term} in name {name}")

	def test_outlets_view(self):
		cls = self.__class__
		cls.driver.get(cls.website)
		outlets = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "OutletTable"))
		ids = self.get_outlet_ids(outlets)
		records = {}
		cls.driver.save_screenshot("./screenshots/test_outlets_view.png")
		# check attributes for each outlet are visible
		# name, format, scale, state all are required
		for i in ids:
			name = f"Outlet_{i}_name"
			scale = f"Outlet_{i}_scale"
			fmt = f"Outlet_{i}_format"
			state = f"Outlet_{i}_state"
			name = cls.driver.find_element(By.ID, name)
			scale = cls.driver.find_element(By.ID, scale)
			fmt = cls.driver.find_element(By.ID, fmt)
			state = cls.driver.find_element(By.ID, state)
			self.assertTrue(name is not None, "Failed to find name cell")
			self.assertTrue(scale is not None, "Failed to find scale cell")
			self.assertTrue(state is not None, "Failed to find state cell")
			self.assertTrue(fmt is not None, "Failed to find format cell")
			name = name.text
			state = state.text
			scale = scale.text
			fmt = fmt.text
			self.assertTrue(name is not None and name != "")
			self.assertTrue(scale is not None and scale != "")
			self.assertTrue(fmt is not None and fmt != "")
			self.assertTrue(state is not None and state != "")

	def test_business_view(self):
		cls = self.__class__
		cls.driver.get(cls.website)
		outlets = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "OutletTable"))
		businesses = cls.driver.find_element(By.ID, "businessPage")
		self.assertTrue(businesses is not None)
		businesses.click()
		entities = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "BusinessTable"))
		self.assertTrue(entities is not None)
		cls.driver.save_screenshot("./screenshots/test_business_view.png")
		rows = entities.find_elements(By.TAG_NAME, "tr")
		self.assertTrue(len(rows) > 0, "Business entity count should be greater than zero")
		ids = []
		for r in rows:
			rowId = r.get_attribute("id")
			if rowId is not None and rowId != "":
				tokens = rowId.split("_")
				entityId = tokens[-1]
				ids.append(entityId)
		for i in ids:
			entityType = f"GenericTableCell_Entity type_{i}"
			name = f"GenericTableCell_Name_{i}"
			name = cls.driver.find_element(By.ID, name)
			entityType = cls.driver.find_element(By.ID, entityType)
			self.assertTrue(name is not None)
			self.assertTrue(entityType is not None)
			name = name.text
			entityType = entityType.text
			self.assertTrue(name is not None and name != "")
			self.assertTrue(entityType is not None and entityType != "")

	def test_filter_outlets_primary_format(self):
		cls = self.__class__
		cls.driver.get(cls.website)
		outlets = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "OutletTable"))
		cls.driver.find_element(By.ID, "primaryFormat").click()
		cls.driver.find_element(By.ID, "primaryFormat_Print").click()
		outlets = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "OutletTable"))
		ids = self.get_outlet_ids(outlets)
		cls.driver.save_screenshot("screenshots/test_filter_outlets_primary_format.png")
		for i in ids:
			fmt = f"Outlet_{i}_format"
			fmt = cls.driver.find_element(By.ID, fmt)
			self.assertTrue(fmt is not None, "Failed to find format cell")
			fmt = fmt.text.strip()
			self.assertEqual(fmt, "Print", "All outlets should a primary format of print")
			
	def test_filter_outlets_scale(self):
		cls = self.__class__
		cls.driver.get(cls.website)
		outlets = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "OutletTable"))
		cls.driver.find_element(By.ID, "scale").click()
		cls.driver.find_element(By.ID, "scale_National").click()
		outlets = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "OutletTable"))
		ids = self.get_outlet_ids(outlets)
		cls.driver.save_screenshot("screenshots/test_filter_outlets_scale.png")
		for i in ids:
			scale = f"Outlet_{i}_scale"
			scale = cls.driver.find_element(By.ID, scale)
			self.assertTrue(scale is not None, "Failed to find format cell")
			scale = scale.text.strip()
			self.assertEqual(scale, "National", "All outlets should a primary format of print")
	
	def test_filter_outlets_subservice(self):
		pass
	
	def test_filter_outlets_broadcast_area(self):
		cls = self.__class__
		cls.driver.get(cls.website)
		outlets = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "OutletTable"))
		cls.driver.find_element(By.ID, "coverage").click()
		baFilter = cls.driver.find_element(By.ID, "broadcast_area_filter")
		ba = cls.terms["broadcast_area_filter"]
		baFilter.send_keys(ba)
		popper = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.CLASS_NAME, "MuiAutocomplete-popper"))
		option = cls.driver.find_element(By.ID, "broadcast_area_filter-option-0")
		ba = option.text.lower()
		option.click()
		outlets = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "OutletTable"))
		ids = self.get_outlet_ids(outlets)
		self.assertTrue(len(ids) > 0, f"There should be at least one outlet in broadcast area {ba}")
		for i in ids:
			entity = f"Outlet_{i}_news_entity"
			entity = cls.driver.find_element(By.ID, entity)
			entity.click()
			button = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "ViewOutlet"))
			button.click()
			container = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "OutletCardContainer"))
			broadcast = cls.driver.find_element(By.ID, "OutletBroadcastAreaCardValue")
			broadcast = broadcast.text.lower().strip()
			self.assertEqual(broadcast, ba, f"Broadcast area should be {ba} but was {broadcast}")
			
	def test_filter_outlets_coverage(self):
		cls = self.__class__
		cls.driver.get(cls.website)
		outlets = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "OutletTable"))
		cls.driver.find_element(By.ID, "coverage").click()
		autocomplete = cls.driver.find_element(By.ID, "coverage_text_filter")
		term = cls.terms["lga"]
		autocomplete.send_keys(term)
		popper = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.CLASS_NAME, "MuiAutocomplete-popper"))
		option = cls.driver.find_element(By.ID, "coverage_text_filter-option-0")
		lga = option.text.lower()
		option.click()
		outlets = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "OutletTable"))
		ids = self.get_outlet_ids(outlets)
		results = []
		for i in ids:
			coverage = cls.driver.find_element(By.ID, f"Outlet_{i}_coverage")
			lgas = coverage.text.strip().lower()
			results.append(lga in lgas)
		self.assertTrue(any(results), "Failed to find any outlet where LGA string {lga} is in coverage")
		cls.driver.save_screenshot("screenshots/test_filter_outlets_coverage.png")
		
	def test_filter_entities_entity_type(self):
		cls = self.__class__
		cls.driver.get(cls.businessWebsite)
		businesses = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "BusinessTable"))
		entityType = cls.driver.find_element(By.ID, "entityType")
		entityType.click()
		option = cls.terms["entityType"]
		checkbox = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, f"entityType_{option}"))
		checkbox.click()
		businesses = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "BusinessTable"))
		ids = self.get_business_ids(businesses)
		self.assertTrue(len(ids) > 0, "There should be at least one business entity")
		option = option.lower()
		for i in ids:
			entityType = f"GenericTableCell_Entity type_{i}"
			entityType = cls.driver.find_element(By.ID, entityType)
			entityType = entityType.text.strip().lower()
			self.assertEqual(entityType, option, f"All visible business entities should be of entity type {option}")
	
	def test_filter_outlets_news_entity(self):
		cls = self.__class__
		cls.driver.get(cls.website)
		outlets = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "OutletTable"))
		cls.driver.find_element(By.ID, "business_accordion").click()
		autocomplete = cls.driver.find_element(By.ID, "news_business_text_filter")
		term = cls.terms["news_entity"]
		autocomplete.send_keys(term)
		popper = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.CLASS_NAME, "MuiAutocomplete-popper"))
		option = cls.driver.find_element(By.ID, "news_business_text_filter-option-0")
		business = option.text.lower()
		option.click()
		outlets = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "OutletTable"))
		ids = self.get_outlet_ids(outlets)
		for i in ids:
			entity = cls.driver.find_element(By.ID, f"Outlet_{i}_news_entity")
			entityName = entity.text.strip().lower()
			self.assertEqual(entityName, business, f"Outlets should have a news entity of {business}, found {entityName}")
	
	def test_view_organisations_table(self):
		cls = self.__class__
		cls.driver.get(cls.website)
		outlets = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "OutletTable"))
		organisations = cls.driver.find_element(By.ID, "organisationPage")
		organisations.click()
		organisations = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "OrganisationsTable"))
		self.assertTrue(organisations is not None, "Organisations table was None")
		ids = self.get_business_ids(organisations)
		self.assertTrue(len(ids) > 0, "There should be at least one organisation record")
		for i in ids:
			name = f"GenericTableCell_Name_{i}"
			organisationType = f"GenericTableCell_Type_{i}"
			shortDescription = f"GenericTableCell_Description_{i}"
			name = cls.driver.find_element(By.ID, name)
			organisationType = cls.driver.find_element(By.ID, organisationType)
			shortDescription = cls.driver.find_element(By.ID, shortDescription)
			self.assertTrue(name is not None)
			self.assertTrue(shortDescription is not None)
			self.assertTrue(organisationType is not None)
			name = name.text.strip()
			organisationType = organisationType.text.strip()
			shortDescription = shortDescription.text.strip()
			self.assertTrue(name is not None and len(name) > 0)
			self.assertTrue(organisationType is not None and len(organisationType) > 0)
			self.assertTrue(organisationType.lower() in cls.ORGANISATION_TYPES, f"Organisation type {organisationType} is not valid.")

	def test_filter_organisations_by_type(self):
		pass
	
	def test_search_organisations_zero_matches(self):
		cls = self.__class__
		cls.driver.get(cls.website)
		outlets = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "OutletTable"))
		organisations = cls.driver.find_element(By.ID, "organisationPage")
		organisations.click()
		organisations = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "OrganisationsTable"))
		self.search_for("DOESNT EXIST")
		organisations = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "OrganisationsTable"))
		ids = self.get_business_ids(organisations)
		self.assertEqual(len(ids), 0, "There should be no matching organisations")

	def test_search_organisations_by_name(self):
		cls = self.__class__
		cls.driver.get(cls.website)
		outlets = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "OutletTable"))
		organisations = cls.driver.find_element(By.ID, "organisationPage")
		organisations.click()
		organisations = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "OrganisationsTable"))
		term = cls.terms["search_organisations"]
		self.search_for(term)
		organisations = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "OrganisationsTable"))
		ids = self.get_business_ids(organisations)
		self.assertEqual(len(ids), 1, "There should be one matching organisation")
		for i in ids:
			name = f"GenericTableCell_Name_{i}"
			name = cls.driver.find_element(By.ID, name)
			name = name.text.strip().lower()
			self.assertTrue(term.lower() in name, f"Organisation records should contain {term} in the name")
		
	def test_view_organisation_cards(self):
		cls = self.__class__
		cls.driver.get(cls.website)
		outlets = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "OutletTable"))
		organisations = cls.driver.find_element(By.ID, "organisationPage")
		organisations.click()
		organisations = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "OrganisationsTable"))
		ids = self.get_business_ids(organisations)
		print("")
		for i in ids:
			print(f"Viewing cards for organisation {i}")
			button = cls.driver.find_element(By.ID, f"View_Organisation_{i}")
			button.click()
			container = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "OrganisationCardContainer"))
			cards = cls.driver.find_elements(By.CLASS_NAME, "EntityCard")
			self.assertTrue(len(cards) > 0, "There should be at least one card present, count was {len(cards)}")
			for c in cards:
				cardId = c.get_attribute("id")
				value = f"{cardId}Value"
				header = f"{cardId}Header"
				# check that the header and value are present
				value = cls.driver.find_element(By.ID, value)
				header = cls.driver.find_element(By.ID, header)
				value = value.text.strip()
				header = header.text.strip()
				self.assertTrue(value != "", "Card values should not be empty")
				self.assertTrue(header != "Card headers should not be empty")
			cls.driver.find_element(By.ID, "OrganisationBack").click()
			organisations = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "OrganisationsTable"))
		
	def get_last_page(self):
		cls = self.__class__
		pages = cls.driver.find_elements(By.CLASS_NAME, "MuiPaginationItem-root")
		numbers = map(lambda x: x.text.strip(), pages)
		possible = [str(x) for x in list(range(1,100))]
		numbers = list(filter(lambda x: x in possible, numbers))
		numbers = [int(x) for x in numbers]
		numbers.sort()
		return numbers[-1]
	
	def view_page(self, n):
		cls = self.__class__
		pages = cls.driver.find_elements(By.CLASS_NAME, "MuiPaginationItem-root")
		for p in pages:
			if str(n) == p.text.strip():
				p.click()
				return
		raise Exception(f"Failed to find page {n}")

	def clear_search(self):
		cls = self.__class__
		search = cls.driver.find_element(By.ID, "ClearSearch")
		search.click()
		
	def search_for(self, expression):
		cls = self.__class__
		search = cls.driver.find_element(By.ID, "search-input")
		search.send_keys(expression)
		search.send_keys(Keys.RETURN)
		
	def test_viewing_all_business_pages(self):
		cls = self.__class__
		cls.driver.get(cls.businessWebsite)
		businesses = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "BusinessTable"))
		i = 1
		pages = self.get_last_page()
		while i <= pages:
			self.view_page(i)
			i += 1
	
	def test_viewing_all_business_entities(self):
		cls = self.__class__
		cls.driver.get(cls.businessWebsite)
		businesses = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "BusinessTable"))
		ids = self.get_business_ids(businesses)
		print("")
		for i in ids:
			print(f"Viewing business with id {i}")
			view = f"View_Business_{i}"
			button = cls.driver.find_element(By.ID, view)
			button.click()
			container = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "BusinessCardContainer"))
			cards = cls.driver.find_elements(By.CLASS_NAME, "EntityCard")
			self.assertTrue(len(cards) > 0, "Failed to find any cards ...")
			back = cls.driver.find_element(By.ID, "BusinessBackButton")
			back.click()
			businesses = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "BusinessTable"))
	
	def test_download_organisations_export(self):
		cls = self.__class__
		cls.driver.get(cls.website)
		outlets = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "OutletTable"))
		organisations = cls.driver.find_element(By.ID, "organisationPage")
		organisations.click()
		organisations = WebDriverWait(cls.driver, timeout=60).until(lambda x: x.find_element(By.ID, "OrganisationsTable"))
		ids = self.get_business_ids(organisations)
		names = []
		for i in ids:
			name = f"GenericTableCell_Name_{i}"
			name = cls.driver.find_element(By.ID, name)
			name = name.text.strip().lower()
			names.append(name)
		cls.driver.find_element(By.ID, "DataMenu").click()
		cls.driver.find_element(By.ID, "downloadExport").click()
		download = "Australian News Index - Industry Bodies.csv"
		destination = None
		timeout = 60
		while True:
			(result, path) = self.does_download_exist(download)
			if result:
					destination = path
					break
			sleep(1)
			timeout -= 1
			if timeout <= 0:
				self.fail("Failed to find export within 60 seconds")
				break
		self.assertTrue(destination is not None)
		with open(destination, "r") as f:
			records = [x.strip().split("|") for x in f.readlines() if len(x.strip()) > 0]
			records = [x for x in records if len(x) > 0]
			records = [records[i] for i in range(1, len(records))] # skip header
			self.assertEqual(len(records), len(ids), "The number of records in the export should match the number of records on screen.") # minus one for header ...
			self.assertTrue(all([x[2].lower() in names for x in records]), "Failed to verify all organisation names match.")
